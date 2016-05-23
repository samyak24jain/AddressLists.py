from openpyxl import load_workbook

wb = load_workbook('New Invitation.xlsx')
readws = wb["ssjs"]
writews = wb["StructuredTable"]


j = 3
for i in range (1,474,6):
	
	#Writing the name
	name = readws.cell('B' + str(i)).value
	writews['A' + str(j)] = name

	#Parsing the firm name
	firmName = readws.cell('B' + str(i+1)).value
	writews['B' + str(j)] = firmName

	#Writing Address Line 1
	addr1 = readws.cell('B' + str(i+2)).value
	writews['C' + str(j)] = addr1

	#Writing Address Line 2
	addr2 = readws.cell('B' + str(i+3)).value
	x = addr2.find("Bangalore")

	if x == -1:
		writews['D' + str(j)] = addr2
	else:
		modAddr2 = addr2[0:x]
		cityAndPincode = addr2[x:]
		writews['D' + str(j)] = modAddr2

	#parsing city and pincode
	pin = ""
	city = cityAndPincode[0:9]
	for a in cityAndPincode:
		if a.isdigit() == True:
			pin += a

	#writing city and pincode
	writews['E' + str(j)] = city
	writews['F' + str(j)] = pin

	#writing phone number
	number = readws.cell('B' + str(i + 4)).value
	for n in number:
		if n.isdigit() == True:
			idx = number.index(n)
			break
	num = number[idx:]
	writews['G' + str(j)] = num

	#Updating j value
	j = j + 1

wb.save('New Invitation.xlsx')